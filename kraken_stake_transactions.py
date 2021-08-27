import time
import datetime
from openpyxl import load_workbook
from kraken_auth import get_kraken_signature,kraken_request
from kraken_market import kraken_price
from secrets import api_key,api_sec


def stake_update(resp):
    # Read most recent staking activity
    top = resp['result'][0]
    top_time_u = top['time']
    top_time = datetime.datetime.fromtimestamp(top_time_u)
    # Read last entry in ledger
    wb = load_workbook(filename = 'Crypto Ledger.xlsx')
    ws = wb.active
    last_row = ws.max_row
    last_time = ws.cell(row = last_row, column = 1).value
    # compare most recent stake entry timestamp to ledger
    if (top_time > last_time):
        next_row = last_row + 1
        asset = str(top['asset'])[:-2]
        asset_index = asset+'USD'
        # get asset price at top_time
        price = kraken_price(asset_index,top_time_u - 3600,60)
        price = price['result'][asset_index][00][4]
        # add new ledger entry
        ws.cell(row = next_row, column = 1).value = top_time
        ws.cell(row = next_row, column = 2).value = 'REWARD'
        ws.cell(row = next_row, column = 3).value = asset
        ws.cell(row = next_row, column = 4).value = top['amount']
        ws.cell(row = next_row, column = 5).value = price
        wb.save('Crypto Ledger.xlsx')
        print('Updated Ledger:',top_time)
        
    else:
        print('No update needed.')

# Construct the request and print the result
resp = kraken_request('/0/private/Staking/Transactions', {
    "nonce": str(int(1000*time.time()))
}, api_key, api_sec)

stake_update(resp.json())
print(resp.json())