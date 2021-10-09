import datetime
from openpyxl import load_workbook
from kraken_auth import kraken_request
from kraken_market import kraken_price

def add_to_ledger(wb,ws,last_row,kraken_time_u,kraken_time,transaction,amount,asset_market,asset):
    next_row = last_row + 1
    # Get asset price at kraken_time
    price = kraken_price(asset_market,kraken_time_u - 3600,60)
    price = price['result'][asset_market][00][4]
    # Add new ledger entry
    ws.cell(row = next_row, column = 1).value = kraken_time
    ws.cell(row = next_row, column = 2).value = transaction
    ws.cell(row = next_row, column = 3).value = asset
    ws.cell(row = next_row, column = 4).value = amount
    ws.cell(row = next_row, column = 5).value = price
    wb.save('Crypto Ledger.xlsx')
    print('Updated Ledger:',kraken_time)


def staking_update(fiat):
    resp = kraken_request('/0/private/Staking/Transactions', {"nonce": str(int(1000*datetime.datetime.utcnow().timestamp()))}).json()
    # Read most recent staking activity
    kraken_time_u = resp['result'][0]['time']
    kraken_time = datetime.datetime.fromtimestamp(kraken_time_u)
    # Read last entry in ledger
    wb = load_workbook(filename = 'Crypto Ledger.xlsx')
    ws = wb.active
    last_row = ws.max_row
    last_time = ws.cell(row = last_row, column = 1).value
    # Compare most recent staking entry timestamp to ledger
    if kraken_time > last_time:
        ahead = 1
        while ahead:
            kraken_time_u = resp['result'][ahead]['time']
            kraken_time = datetime.datetime.fromtimestamp(kraken_time_u)
            if kraken_time > last_time:
                ahead = ahead + 1

            else:
                ahead = ahead - 1
                # Set columns
                kraken_time_u = resp['result'][ahead]['time']
                kraken_time = datetime.datetime.fromtimestamp(kraken_time_u)
                transaction = resp['result'][ahead]['type'].upper()
                asset = str(resp['result'][ahead]['asset'])[:-2]
                asset_market = asset+fiat
                amount = resp['result'][ahead]['amount']

                add_to_ledger(wb,ws,last_row,kraken_time_u,kraken_time,transaction,amount,asset_market,asset)
                # Update last row and time for next iteration
                last_row = ws.max_row
                last_time = ws.cell(row = last_row, column = 1).value

    else:
        print('No update needed.')

# Check for stake rewards and record them in the ledger
# A fiat currency string is used to find the approximate price of an asset at the time of the staking transaction
staking_update('USD')
